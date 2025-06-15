import os  # ファイル操作用
import pdfplumber  # PDF処理
import pandas as pd  # データ処理
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

INPUT_DIR = "input"
OUTPUT_DIR = "output"

def ensure_dirs():
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def normalize_text(text):
    if not text:
        return ""
    text = text.replace('\u3000', ' ')
    text = re.sub(r'\s+', '', text)
    return text

def parse_timetable(table):
    cleaned_rows = [row for row in table if sum(1 for cell in row if cell) >= 2]

    if not cleaned_rows or len(cleaned_rows) < 2:
        return pd.DataFrame()

    header = cleaned_rows[0]
    data_rows = cleaned_rows[1:]

    row_keywords = ["備考", "注意"]
    filtered_rows = []
    for row in data_rows:
        normalized_cells = [normalize_text(str(cell)) for cell in row]
        if not any(any(kw in cell for kw in row_keywords) for cell in normalized_cells):
            filtered_rows.append(row)

    clean_header = [str(col).strip().replace('\n', ' ') if col else '' for col in header]

    col_keywords = ["年次", "担当", "備考", "曜"]
    normalized_header = [normalize_text(col) for col in clean_header]
    col_indices_to_keep = [i for i, col in enumerate(normalized_header) if not any(kw in col for kw in col_keywords)]

    adjusted_rows = [row + [''] * (len(clean_header) - len(row)) for row in filtered_rows]

    df = pd.DataFrame(adjusted_rows, columns=clean_header)
    df = df.convert_dtypes()
    return df

def process_pdf(file_path):
    with pdfplumber.open(file_path) as pdf:
        all_data = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table:
                    df = parse_timetable(table)
                    all_data.append(df)
        return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()

def autofit_excel_columns(file_path): 
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)

def format_excel_cells(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )
                cell.font = Font(name="MS PGothic", size=10)
        sheet.sheet_view.zoomScale = 90
    wb.save(file_path)

def main():
    ensure_dirs()
    for file_name in os.listdir(INPUT_DIR):
        if file_name.lower().endswith(".pdf"):
            input_path = os.path.join(INPUT_DIR, file_name)
            print(f"Processing {file_name}...")
            df = process_pdf(input_path)
            if not df.empty:
                output_name = os.path.splitext(file_name)[0] + ".xlsx"
                output_path = os.path.join(OUTPUT_DIR, output_name)
                df.to_excel(output_path, index=False)
                autofit_excel_columns(output_path)
                format_excel_cells(output_path)
                print(f"Saved to {output_path}")
            else:
                print(f"No data found in {file_name}")

if __name__ == "__main__":
    main()
